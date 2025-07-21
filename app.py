from flask import Flask, render_template, request, redirect, url_for, jsonify
import sqlite3
from datetime import datetime
from collections import defaultdict, OrderedDict
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)

# Fungsi untuk menginisialisasi database
def init_db():
    conn = sqlite3.connect('survey.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS surveys (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            age INTEGER NOT NULL,
            phone TEXT NOT NULL,
            alamat TEXT NOT NULL,
            kecamatan TEXT NOT NULL,
            kabupaten TEXT NOT NULL,
            accident_date TEXT NOT NULL,
            location TEXT NOT NULL,
            vehicle TEXT NOT NULL,
            tax_status TEXT NOT NULL,
            hospital TEXT NOT NULL, 
            q1 TEXT NOT NULL,
            q2 TEXT NOT NULL,
            q3 TEXT NOT NULL,
            q4 TEXT NOT NULL,
            q5 TEXT NOT NULL,
            q6 TEXT NOT NULL,
            q7 TEXT NOT NULL,
            q8 TEXT NOT NULL,
            q9 TEXT NOT NULL,
            q9_reason TEXT,
            q10 TEXT NOT NULL,
            q10_reason TEXT,
            suggestion TEXT,
            statement BOOLEAN NOT NULL,
            understanding BOOLEAN NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        )
    ''')
    conn.commit()
    conn.close()

# Route untuk halaman survei
@app.route('/')
def index():
    return render_template('home.html')

@app.route('/survey')
def survey():
    return render_template('survey.html')

# Route untuk pengiriman survei
@app.route('/submit', methods=['POST'])
def submit():
    if request.method == 'POST':
        # Ambil data dari form
        name = request.form['name']
        age = request.form['age']
        phone = request.form['phone']
        alamat = request.form['alamat']
        kecamatan = request.form['kecamatan']
        kabupaten = request.form['kabupaten']
        accident_date = request.form['accident-date']
        location = request.form['location']
        vehicle = request.form['vehicle']
        tax_status = request.form['tax-status']
        hospital = request.form['hospital'] 
        q1 = request.form['q1']
        q2 = request.form['q2']
        q3 = request.form['q3']
        q4 = request.form['q4']
        q5 = request.form['q5']
        q6 = request.form['q6']
        q7 = request.form['q7']
        q8 = request.form['q8']
        q9 = request.form['q9']
        q9_reason = request.form.get('q9-reason', '')
        q10 = request.form['q10']
        q10_reason = request.form.get('q10-reason', '')
        suggestion = request.form.get('suggestion', '')
        statement = 'statement' in request.form
        understanding = 'understanding' in request.form

        # Simpan data ke database
        conn = sqlite3.connect('survey.db')
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO surveys (
                name, age, phone, alamat, accident_date, location, vehicle, tax_status,
                hospital, kecamatan, kabupaten, q1, q2, q3, q4, q5, q6, q7, q8, q9,
                q9_reason, q10, q10_reason, suggestion, statement, understanding
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            name, age, phone, alamat, accident_date, location, vehicle, tax_status,
            hospital, kecamatan, kabupaten, q1, q2, q3, q4, q5, q6, q7, q8, q9,
            q9_reason, q10, q10_reason, suggestion, statement, understanding
        ))
        conn.commit()
        conn.close()

        return render_template('thank_you.html')


# Nilai untuk setiap jawaban
nilai_mapping = {
    'q1': {'Ada': 4, 'Tidak': 1},
    'q4': {'Seluruhnya dibantu oleh Petugas Jasa Raharja': 4, 'Seluruhnya dibantu oleh Petugas Polisi atau Pihak ke tiga': 2, 'Diurus sendiri oleh korban ahli waris korban': 1},
    'q5': {'Tidak Pernah': 4, '1 Kali': 3, '2 Kali': 2, 'Lebih dari 2 Kali': 1},
    'q6': {'Tidak Pernah': 4, '1 Kali': 3, '2 Kali': 2, 'Lebih dari 2 Kali': 1},
    'q7': {'Wajar': 4, 'Tidak Wajar': 1},
    'q8': {'Tidak Ada': 4, 'Ada': 1},
    'q9': {'Sangat Memuaskan': 4, 'Memuaskan': 3, 'Kurang Memuaskan': 2, 'Tidak Memuaskan': 1} 
}
# Alias jawaban agar sesuai dengan urutan manual
alias_mapping = {
    'q2': {
        'Petugas Jasa Raharja': 'Petugas Jasa Raharja',
        'Petugas Rumah Sakit': 'Petugas Rumah Sakit',
        'Media Massa': 'Media Massa',
        'Petugas Kepolisian': 'Petugas Kepolisian',
        'Lainnya': 'Orang yang pernah menerima santunan/masyarakat umum',
        # tambahkan jika ada variasi lain
    },
    'q3': {
        'Ahli Waris': 'Ahli Waris Korban',
        'Korban': 'Korban',
        'Keluarga': 'Keluarga Korban',
        'Pihak Ketiga': 'Pihak Ketiga'
    },
    'q4': {
        'Petugas Jasa Raharja': 'Seluruhnya dibantu oleh Petugas Jasa Raharja',
        'Petugas Polisi': 'Seluruhnya dibantu oleh Petugas Polisi atau Pihak ke tiga',
        'Sendiri': 'Diurus sendiri oleh korban ahli waris korban'
    }
}
# Urutan manual untuk pertanyaan non-skor
manual_order = {
    'q2': ['Petugas Jasa Raharja', 'Petugas Rumah Sakit', 'Petugas Kepolisian', 'Media Massa', 'Orang yang pernah menerima santunan/masyarakat umum'],
    'q3': ['Korban', 'Ahli Waris Korban', 'Keluarga Korban', 'Pihak Ketiga'],
    'q4': ['Seluruhnya dibantu oleh Petugas Jasa Raharja', 'Seluruhnya dibantu oleh Petugas Polisi atau Pihak ke tiga', 'Diurus sendiri oleh korban ahli waris korban'],
    'q10': ['Sangat Memuaskan', 'Memuaskan', 'Kurang Memuaskan', 'Tidak Memuaskan']
}
# Label semua pertanyaan
pertanyaan_label = {
    'q1': '1. Ada Petugas JR datang memberi informasi?',
    'q2': '2. Darimana korban mengetahui dalam tanggungan Jasa Raharja?',
    'q3': '3. Dalam pengurusan dokumen persyaratan santunan diurus oleh?',
    'q4': '4. Pelayanan terpadu dalam pengurusan persyaratan santunan, Bagaimana pengurusan surat-surat tersebut?',
    'q5': '5. Penerima santunan telah mendatangi kantor Jasa Raharja?',
    'q6': '6. Penerima santunan dalam pengurusan surat-surat telah mendatangi instansi terkait selain Jasa Raharja?',
    'q7': '7. Bagaimana tingkat kewajaran biaya rawatan dapat disimpulkan?',
    'q8': '8. Adakah biaya yang diberikan kepada Petugas Jasa Raharja dalam pengurusan Santunan?',
    'q9': '9. Bagaimana Pelayanan Petugas Jasa Raharja Secara Umum?',
    'q10': '10. Bagaimana Pelayanan Rumah Sakit Secara Umum?'
}
@app.route('/laporan')
def laporan():
    import sqlite3
    from collections import defaultdict, OrderedDict

    conn = sqlite3.connect('survey.db')
    cursor = conn.cursor()

    semua_pertanyaan = list(pertanyaan_label.keys())
    query = f"SELECT {', '.join(semua_pertanyaan)} FROM surveys"
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()

    # Inisialisasi rekap mentah
    rekap_mentah = {q: defaultdict(int) for q in semua_pertanyaan}

    for row in rows:
        for idx, q in enumerate(semua_pertanyaan):
            jawaban = row[idx]
            rekap_mentah[q][jawaban] += 1

    total_skor = 0
    rekap_terurut = {}

    for q in semua_pertanyaan:
        jawaban_dict = defaultdict(int)

        # Gabungkan dan alias-kan jawaban
        for jawaban_asli, count in rekap_mentah[q].items():
            jawaban_bersih = alias_mapping.get(q, {}).get(jawaban_asli, jawaban_asli)
            jawaban_dict[jawaban_bersih] += count

        # Hitung skor (dihitung dari jawaban bersih)
        if q in nilai_mapping:
            for jawaban, jumlah in jawaban_dict.items():
                nilai = nilai_mapping[q].get(jawaban)
                if nilai is not None:
                    total_skor += jumlah * nilai

        # Urutkan jawaban
        if q in manual_order:
            ordered = sorted(
                jawaban_dict.items(),
                key=lambda item: manual_order[q].index(item[0]) if item[0] in manual_order[q] else 999
            )
        elif q in nilai_mapping:
            ordered = sorted(
                jawaban_dict.items(),
                key=lambda item: nilai_mapping[q].get(item[0], 0),
                reverse=True
            )
        else:
            ordered = sorted(jawaban_dict.items())

        rekap_terurut[q] = OrderedDict(ordered)

    jumlah_pertanyaan_skor = len(nilai_mapping)
    total_responden = len(rows)
    maks_skor = total_responden * jumlah_pertanyaan_skor * 4
    nilai_survei = (total_skor / maks_skor) * 100 if maks_skor > 0 else 0

    # Penilaian predikat
    if nilai_survei <= 70:
        predikat = "Kurang Memuaskan"
    elif nilai_survei <= 80:
        predikat = "Cukup Memuaskan"
    elif nilai_survei <= 90:
        predikat = "Memuaskan"
    else:
        predikat = "Sangat Memuaskan"

    return render_template(
        'laporan.html',
        rekap=rekap_terurut,
        nilai_mapping=nilai_mapping,
        pertanyaan_label=pertanyaan_label,
        total_responden=total_responden,
        total_skor=total_skor,
        jumlah_pertanyaan_skor=jumlah_pertanyaan_skor,
        nilai_survei=round(nilai_survei, 2),
        predikat=predikat
    )


@app.route("/export_excel")
def export_excel():
    from flask import request, send_file
    import sqlite3
    import pandas as pd
    conn = sqlite3.connect('survey.db')
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = """
    SELECT * FROM surveys
    WHERE date(created_at) BETWEEN ? AND ?
    """
    df = pd.read_sql_query(query, conn, params=(start_date, end_date))

    output_path = "filtered_survey.xlsx"
    df.to_excel(output_path, index=False)
    return send_file(output_path, as_attachment=True)

@app.route('/laporan/excel')
def export_laporan_excel():
    import sqlite3
    import pandas as pd
    from collections import defaultdict, OrderedDict
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # ==== DATABASE ====
    conn = sqlite3.connect('survey.db')
    cursor = conn.cursor()
    semua_pertanyaan = list(pertanyaan_label.keys())
    query = f"SELECT {', '.join(semua_pertanyaan)} FROM surveys"
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()

    # ==== HITUNG REKAP ====
    rekap_mentah = {q: defaultdict(int) for q in semua_pertanyaan}
    for row in rows:
        for idx, q in enumerate(semua_pertanyaan):
            jawaban = row[idx]
            rekap_mentah[q][jawaban] += 1

    rekap_terurut = {}
    total_skor = 0
    total_responden = len(rows)

    for q in semua_pertanyaan:
        jawaban_dict = defaultdict(int)
        for jawaban_asli, count in rekap_mentah[q].items():
            jawaban_bersih = alias_mapping.get(q, {}).get(jawaban_asli, jawaban_asli)
            jawaban_dict[jawaban_bersih] += count

        if q in nilai_mapping:
            for jawaban, jumlah in jawaban_dict.items():
                nilai = nilai_mapping[q].get(jawaban)
                if nilai:
                    total_skor += jumlah * nilai

        if q in manual_order:
            ordered = sorted(jawaban_dict.items(), key=lambda item: manual_order[q].index(item[0]) if item[0] in manual_order[q] else 999)
        elif q in nilai_mapping:
            ordered = sorted(jawaban_dict.items(), key=lambda item: nilai_mapping[q].get(item[0], 0), reverse=True)
        else:
            ordered = sorted(jawaban_dict.items())

        rekap_terurut[q] = OrderedDict(ordered)

    jumlah_pertanyaan_skor = len(nilai_mapping)
    maks_skor = total_responden * jumlah_pertanyaan_skor * 4
    nilai_survei = (total_skor / maks_skor) * 100 if maks_skor > 0 else 0

    if nilai_survei <= 70:
        predikat = "Kurang Memuaskan"
    elif nilai_survei <= 80:
        predikat = "Cukup Memuaskan"
    elif nilai_survei <= 90:
        predikat = "Memuaskan"
    else:
        predikat = "Sangat Memuaskan"

    # ==== STYLING ====
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Survei"

    # Gaya
    bold_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    for q in semua_pertanyaan:
        # Judul pertanyaan
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=1, value=pertanyaan_label[q]).font = bold_font
        ws.cell(row=current_row, column=1).alignment = center
        current_row += 1

        # Header tabel
        headers = ['Jawaban', 'Jumlah Responden', 'Nilai', 'Skor Responden']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = bold_font
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border
        current_row += 1

        # Isi data
        for jawaban, jumlah in rekap_terurut[q].items():
            nilai = nilai_mapping[q].get(jawaban) if q in nilai_mapping else None
            skor = jumlah * nilai if nilai else None

            ws.cell(row=current_row, column=1, value=jawaban)
            ws.cell(row=current_row, column=2, value=jumlah)
            ws.cell(row=current_row, column=3, value=nilai if nilai else '-')
            ws.cell(row=current_row, column=4, value=skor if skor else '-')

            for col in range(1, 5):
                ws.cell(row=current_row, column=col).alignment = center
                ws.cell(row=current_row, column=col).border = border
            current_row += 1

        current_row += 2  # Spasi antar pertanyaan

    # ==== RINGKASAN ====
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.cell(row=current_row, column=1, value="RINGKASAN").font = Font(bold=True, color="000000")
    ws.cell(row=current_row, column=1).alignment = center
    current_row += 1

    ringkasan = [
        ['Total Skor Responden Real', total_skor],
        ['Total Skor Responden Maksimal', maks_skor],
        ['Nilai Survei (%)', f"{round(nilai_survei, 2)}%"],
        ['Predikat Survei', predikat],
    ]

    for keterangan, nilai in ringkasan:
        ws.cell(row=current_row, column=1, value=keterangan).font = bold_font
        ws.cell(row=current_row, column=2, value=nilai)
        ws.cell(row=current_row, column=1).alignment = center
        ws.cell(row=current_row, column=2).alignment = center

        for col in range(1, 3):
            ws.cell(row=current_row, column=col).border = border
        current_row += 1

    # Lebar kolom
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20

    # Simpan ke BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='Laporan_Survei_JR.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/report')
def report():
    conn = sqlite3.connect('survey.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM surveys')
    rows = cursor.fetchall()
    conn.close()

    columns = [description[0] for description in cursor.description]
    data = [dict(zip(columns, row)) for row in rows]
    print(data)
    return render_template('report.html', data=data) 

@app.route('/report/<int:id>')
def detail(id):
    conn = sqlite3.connect('survey.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM surveys WHERE id = ?', (id,))
    row = cursor.fetchone()
    conn.close() 
    columns = [description[0] for description in cursor.description]
    data = dict(zip(columns, row))
    print(row)
    return render_template('report-detail.html', data=data)

@app.route('/report/delete/<int:id>', methods=['POST'])
def delete(id):
    conn = sqlite3.connect('survey.db')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM surveys WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('report'))

@app.route('/report/export')
def export():
    conn = sqlite3.connect('survey.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM surveys')
    rows = cursor.fetchall()
    conn.close()
    # Implement export functionality here
    return "Data exported successfully"
 

@app.route("/dashboard")
def dashboard():
    conn = sqlite3.connect("survey.db")
    cursor = conn.cursor()

    # Total Responden
    cursor.execute("SELECT COUNT(*) FROM surveys")
    total_responses = cursor.fetchone()[0]

    # Usia Rata-rata
    cursor.execute("SELECT AVG(age) FROM surveys")
    avg_age = int(cursor.fetchone()[0] or 0)

    # Rata-rata Kepuasan (skor 1–5)
    # cursor.execute("""
    #     SELECT AVG(CASE q9
    #         WHEN 'Sangat Memuaskan' THEN 100
    #         WHEN 'Memuaskan' THEN 80
    #         WHEN 'Kurang Memuaskan' THEN 60
    #         WHEN 'Tidak Memuaskan' THEN 40
    #     END) FROM surveys
    # """)
    # avg_satisfaction = round(cursor.fetchone()[0] or 0, 2)

    # Hitung nilai_survei dari q9 (Sangat Memuaskan → 4, ... Tidak Memuaskan → 1)

    semua_pertanyaan = list(pertanyaan_label.keys())
    query = f"SELECT {', '.join(semua_pertanyaan)} FROM surveys"
    cursor.execute(query)
    rows = cursor.fetchall() 

    # Inisialisasi rekap mentah
    rekap_mentah = {q: defaultdict(int) for q in semua_pertanyaan}

    for row in rows:
        for idx, q in enumerate(semua_pertanyaan):
            jawaban = row[idx]
            rekap_mentah[q][jawaban] += 1

    total_skor = 0
    rekap_terurut = {}

    for q in semua_pertanyaan:
        jawaban_dict = defaultdict(int)

        # Gabungkan dan alias-kan jawaban
        for jawaban_asli, count in rekap_mentah[q].items():
            jawaban_bersih = alias_mapping.get(q, {}).get(jawaban_asli, jawaban_asli)
            jawaban_dict[jawaban_bersih] += count

        # Hitung skor (dihitung dari jawaban bersih)
        if q in nilai_mapping:
            for jawaban, jumlah in jawaban_dict.items():
                nilai = nilai_mapping[q].get(jawaban)
                if nilai is not None:
                    total_skor += jumlah * nilai

        # Urutkan jawaban
        if q in manual_order:
            ordered = sorted(
                jawaban_dict.items(),
                key=lambda item: manual_order[q].index(item[0]) if item[0] in manual_order[q] else 999
            )
        elif q in nilai_mapping:
            ordered = sorted(
                jawaban_dict.items(),
                key=lambda item: nilai_mapping[q].get(item[0], 0),
                reverse=True
            )
        else:
            ordered = sorted(jawaban_dict.items())

        rekap_terurut[q] = OrderedDict(ordered)

    jumlah_pertanyaan_skor = len(nilai_mapping)
    total_responden = len(rows)
    maks_skor = total_responden * jumlah_pertanyaan_skor * 4
    avg_satisfaction = round((total_skor / maks_skor) * 100 if maks_skor > 0 else 0 or 0, 2)


    # Tanggal terakhir diisi
    cursor.execute("SELECT MAX(accident_date) FROM surveys")
    last_date = cursor.fetchone()[0]
    response_rate = "-"
    if last_date:
        y, m, d = last_date.split("-")
        bulan_indo = {
            '01': 'Januari', '02': 'Februari', '03': 'Maret', '04': 'April',
            '05': 'Mei', '06': 'Juni', '07': 'Juli', '08': 'Agustus',
            '09': 'September', '10': 'Oktober', '11': 'November', '12': 'Desember'
        }
        response_rate = f"{int(d)} {bulan_indo[m]} {y}"

    # Distribusi Kepuasan
    cursor.execute("SELECT q9, COUNT(*) FROM surveys GROUP BY q9")
    satisfaction_data = dict(cursor.fetchall())

    # Status Pajak
    cursor.execute("SELECT tax_status, COUNT(*) FROM surveys GROUP BY tax_status")
    tax_status_data = dict(cursor.fetchall())

    # Kabupaten
    cursor.execute("""
        SELECT kabupaten, COUNT(*) FROM surveys
        GROUP BY kabupaten ORDER BY COUNT(*) DESC LIMIT 5
    """)
    kabupaten_data = dict(cursor.fetchall())

    # Rumah Sakit
    cursor.execute("""
        SELECT hospital, COUNT(*) FROM surveys
        GROUP BY hospital ORDER BY COUNT(*) DESC LIMIT 5
    """)
    hospital_data = dict(cursor.fetchall())

    # Trend mingguan
    cursor.execute("""
        SELECT strftime('%Y-%W', accident_date) AS minggu, COUNT(*)
        FROM surveys GROUP BY minggu ORDER BY minggu
    """)
    trend_data = cursor.fetchall()

    # Ambil saran terbaru
    cursor.execute("""
        SELECT suggestion, accident_date FROM surveys
        WHERE suggestion IS NOT NULL AND suggestion != ''
        ORDER BY accident_date DESC LIMIT 5
    """)
    raw_suggestions = cursor.fetchall()
    suggestions = []
    for text, date in raw_suggestions:
        if date:
            y, m, d = date.split("-")
            tanggal = f"{int(d)} {bulan_indo[m]} {y}"
        else:
            tanggal = "-"
        suggestions.append({"text": text, "date": tanggal})

    conn.close()

    return render_template("dashboard.html",
        total_responses=total_responses,
        avg_age=avg_age,
        avg_satisfaction=avg_satisfaction,
        response_rate=response_rate,
        satisfaction_data=satisfaction_data,
        tax_status_data=tax_status_data,
        kabupaten_data=kabupaten_data,
        hospital_data=hospital_data,
        trend_data=trend_data,
        suggestions=suggestions
    )

@app.route('/dashboarde')
def dashboarde():    
    bulan_indo = {
        '01': 'Januari', '02': 'Februari', '03': 'Maret',
        '04': 'April', '05': 'Mei', '06': 'Juni',
        '07': 'Juli', '08': 'Agustus', '09': 'September',
        '10': 'Oktober', '11': 'November', '12': 'Desember'
    }
    conn = sqlite3.connect('survey.db')
    cursor = conn.cursor()
    
    # Total responses
    cursor.execute('SELECT COUNT(*) FROM surveys')
    total_responses = cursor.fetchone()[0]
    
    # Average age
    cursor.execute('SELECT AVG(age) FROM surveys')
    avg_age = cursor.fetchone()[0]
    avg_age = int(avg_age)
    # Average satisfaction
    cursor.execute('''
        SELECT AVG(CAST(
            CASE q9
                WHEN 'Sangat Memuaskan' THEN 5
                WHEN 'Memuaskan' THEN 4
                WHEN 'Kurang Memuaskan' THEN 2
                WHEN 'Tidak Memuaskan' THEN 1
            END AS INTEGER)
        ) FROM surveys
    ''')
    avg_satisfaction = cursor.fetchone()[0]
    avg_satisfaction = round(avg_satisfaction, 2)
    
    # Last response date
    cursor.execute('SELECT MAX(accident_date) FROM surveys')
    last_response_date = cursor.fetchone()[0]
    if last_response_date is None:
        response_rate = "Belum ada data"
    else:
        from datetime import datetime
        bulan_indo = {
            '01': 'Januari', '02': 'Februari', '03': 'Maret',
            '04': 'April', '05': 'Mei', '06': 'Juni',
            '07': 'Juli', '08': 'Agustus', '09': 'September',
            '10': 'Oktober', '11': 'November', '12': 'Desember'
        }

        tahun, bulan, hari = last_response_date.split('-')
        response_rate = f"{int(hari)} {bulan_indo[bulan]} {tahun}"
    
    # Satisfaction data
    cursor.execute('''
        SELECT q9, COUNT(*) 
        FROM surveys 
        GROUP BY q9
    ''')
    satisfaction_data = cursor.fetchall()
    
    # Tax status data
    cursor.execute('''
        SELECT tax_status, COUNT(*) 
        FROM surveys 
        GROUP BY tax_status
    ''')
    tax_status_data = cursor.fetchall()
    
    # Response trend data
    cursor.execute('''
        SELECT strftime('%Y-%W', accident_date) AS week, COUNT(*) 
        FROM surveys 
        GROUP BY week
    ''')
    response_trend_data = cursor.fetchall()
    
    # Hospital data
    cursor.execute('''
        SELECT hospital, COUNT(*) 
        FROM surveys 
        GROUP BY hospital
        ORDER BY COUNT(*) DESC
        LIMIT 5
    ''')
    hospital_data = cursor.fetchall()
    
    # Kabupaten data
    cursor.execute('''
        SELECT kabupaten, COUNT(*) 
        FROM surveys 
        GROUP BY kabupaten
        ORDER BY COUNT(*) DESC
        LIMIT 5
    ''')
    kabupaten_data = cursor.fetchall()
    
    # Suggestions
    cursor.execute('''
        SELECT suggestion, accident_date 
        FROM surveys 
        WHERE suggestion IS NOT NULL 
        ORDER BY accident_date DESC 
        LIMIT 5
    ''')
    suggestions = cursor.fetchall()
    columns = [description[0] for description in cursor.description]
    suggestions_data = [dict(zip(columns, row)) for row in suggestions]

    for item in suggestions_data:
        tahun, bulan, hari = item['accident_date'].split('-')
        item['accident_date'] = f"{int(hari)} {bulan_indo[bulan]} {tahun}"
    conn.close()
    print(columns)
    
    return render_template(
        'dashboard.html', 
        total_responses=total_responses,
        avg_age=avg_age,
        avg_satisfaction=avg_satisfaction,
        response_rate=response_rate,
        suggestions=suggestions_data
    )
 

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)